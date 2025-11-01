from time import sleep
def limpar_tela():
    import os
    os.system('cls' if os.name=='nt'else 'clear')

def linha(tam=50):
    print('-'*tam)

def cabecalho(msg):
    linha()
    print(msg)
    linha()

def balanco():
    from openpyxl import Workbook, load_workbook
    from time import sleep
    extrato = [[], []]
    while True:
        cabecalho('Escolha o tipo de operação')
        while True:
            tipo = str(input("Tipo de Operação [D ou C]: ")).strip().upper()[0]
            if tipo not in 'DC':
                print('Digite Tipo de Operação [D] ou [C]')
            else:
                break
        if tipo in 'D':
            extrato[1].append(float(input('Digite o Valor do Débito: R$ ')))
        elif tipo in 'C':
            extrato[0].append(float(input('Digite o valor do Crédito: R$ ')))
        resp = str(input('Quer Continuar ?(S/N) ')).strip().upper()[0]
        if resp in 'Nn':
            break
    print('-'*50)
    limpar_tela()
    cabecalho('Créditos Digitados: ')
    for i, credito in enumerate(extrato[0]):
        print(f'{i} ==> R$ {credito:.2f}')
    sleep(0.3)
    cabecalho('Débitos Digitados: ')
    for j, debito in enumerate(extrato[1]):
        print(f'{j} ==> R$ {debito:.2f}')
    print('-'*50)
    sleep(0.3)
    datainicial = str(input('Digite data (dd/mm/aaaa): ')).strip()
    if len(extrato[0]) == 0:
        entradas = 0
    else:
        entradas = sum(extrato[0])
    if len(extrato[1]) == 0:
        saidas = 0
    else:
        saidas = sum(extrato[1])
    balanco = entradas - saidas
    print(f'{saidas} -> {entradas} -> {balanco}')
    wb = load_workbook(r"G:\Meu Drive\despesas\balanco.xlsx")
    ws = wb.active
    linha = ws.max_row
    ws.cell(row=linha + 1, column=1, value=datainicial)
    ws.cell(row=linha + 1, column=2, value=entradas)
    ws.cell(row=linha + 1, column=3, value=saidas)
    ws.cell(row=linha + 1, column=4, value=balanco)
    try:
        wb.save(r"G:\Meu Drive\despesas\balanco.xlsx")
    except:
        print('Erro ao gravar arquivo.')
    else:
        print('Registro adicionado com sucesso !')
    return
def despesas(msg="WILLIAM"):
    from openpyxl import Workbook, load_workbook
    limpar_tela()
    cabecalho('Entre com os lançamentos de Débitos e Créditos ')
    while True:
        descricao = str(input('Digite a descrição do lançamento: ')).strip().upper()
        data = str(input('Digite a Data do lançamento (dd/mm/aaaa): ')).strip()
        valor = float(input('Digite o valor do lançamento: R$ '))
        status = str(input('Seu lançamento é um Crédito ou um Débito ? ')).strip().upper()
        opcao = str(input('Deseja gravar na planilha ?(S/N) ')).strip().upper()[0]
        if opcao in 'S':
            wb = load_workbook(r"G:\Meu Drive\despesas\despesas.xlsx")
            ws = wb.active
            ws = wb[msg]
            linha = ws.max_row
            ws.cell(row=linha + 1, column=1, value=data)
            ws.cell(row=linha + 1, column=2, value=descricao)
            ws.cell(row=linha + 1, column=3, value=valor)
            ws.cell(row=linha + 1, column=4, value=status)
            try:
                wb.save(r"G:\Meu Drive\despesas\despesas.xlsx")
            except:
                print('Erro ao gravar arquivo.')
            else:
                print('Registro adicionado com sucesso !')
        else:
            print('Registro não gravado na planilha.')
        resp = str(input('Quer fazer novo lançamento ? (S/N) ')).strip().upper()[0]
        if resp in 'N':
            cabecalho('Obrigado pela Colaboração !')
            sleep(0.5)
            limpar_tela()
            break
    return